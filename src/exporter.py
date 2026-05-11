from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

MAX_COL_WIDTH = 60  # cap autosize so a giant article body doesn't blow out the column


def to_excel(df: pd.DataFrame, output_path: Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    sheet_name = "coal_articles"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]

        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"

        for col_idx, column in enumerate(df.columns, start=1):
            header_len = len(str(column))
            sample = df[column].astype(str).head(50)
            max_data_len = sample.map(len).max() if not sample.empty else 0
            width = min(MAX_COL_WIDTH, max(header_len, int(max_data_len)) + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    log.info("Wrote Excel file: %s (%d rows)", output_path, len(df))
    return output_path.resolve()
